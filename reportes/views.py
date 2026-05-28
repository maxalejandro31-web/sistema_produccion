from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from inventario.models import MateriaPrima, MovimientoMP, Cliente
from produccion.models import OrdenProduccion, DetalleSlitter
from materia_terminada.models import ProductoTerminado


# ── Helpers de estilo ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F2D3D")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _apply_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
    ws.row_dimensions[1].height = 22


def _apply_rows(ws):
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if i % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill
            cell.border    = BORDER
            cell.alignment = LEFT


def _autowidth(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)


def _make_response(wb, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _fmt(val):
    return val if val is not None else ""


# ── Vista principal ────────────────────────────────────────────────────────────

@login_required
def reportes_index(request):
    clientes = Cliente.objects.filter(activo=True).order_by("nombre")
    operadores = None
    try:
        from produccion.models import Operador
        operadores = Operador.objects.filter(activo=True).order_by("nombre")
    except Exception:
        pass
    return render(request, "reportes/index.html", {
        "clientes": clientes,
        "operadores": operadores,
    })


# ── 1. Inventario Materia Prima ────────────────────────────────────────────────

@login_required
def reporte_inventario_mp(request):
    qs = MateriaPrima.objects.select_related("cliente").all()

    cliente_id = request.GET.get("cliente")
    estado     = request.GET.get("estado")
    tipo_mp    = request.GET.get("tipo_mp")
    origen_mp  = request.GET.get("origen_mp")

    if cliente_id:
        qs = qs.filter(cliente_id=cliente_id)
    if estado:
        qs = qs.filter(estado=estado)
    if tipo_mp:
        qs = qs.filter(tipo_mp=tipo_mp)
    if origen_mp:
        qs = qs.filter(origen_mp=origen_mp)

    qs = qs.order_by("-fecha_entrada", "numero_mp")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario MP"

    headers = [
        "N° MP", "Tipo", "Cliente", "Origen", "Lote", "Código", "Descripción",
        "Material", "Grado", "Acabado", "Espesor", "Unidad Espesor",
        "Ancho (mm)", "Largo (mm)", "Peso Inicial (kg)", "Peso Restante (kg)",
        "Diam. Interior", "Diam. Exterior",
        "Proveedor", "Ubicación", "Estado", "Fecha Entrada",
        "Días en Fábrica", "Estatus Cobro", "Meses a Cobrar",
        "Observaciones",
    ]
    _apply_header(ws, headers)

    ESTATUS_LABEL = {
        'libre':    'Libre',
        'por_vencer': 'Por vencer',
        'vencido':  'Cobro activo',
        'sin_fecha': 'Sin fecha',
    }

    for mp in qs:
        ws.append([
            _fmt(mp.numero_mp),
            _fmt(mp.tipo_mp),
            _fmt(mp.cliente.nombre if mp.cliente else ""),
            _fmt(mp.origen_mp),
            _fmt(mp.lote),
            _fmt(mp.codigo),
            _fmt(mp.descripcion),
            _fmt(mp.material),
            _fmt(mp.grado),
            _fmt(mp.acabado),
            float(mp.espesor_valor) if mp.espesor_valor else "",
            _fmt(mp.unidad_espesor),
            float(mp.ancho) if mp.ancho else "",
            float(mp.largo) if mp.largo else "",
            float(mp.peso) if mp.peso else "",
            float(mp.peso_restante) if mp.peso_restante else "",
            float(mp.diametro_interior) if mp.diametro_interior else "",
            float(mp.diametro_exterior) if mp.diametro_exterior else "",
            _fmt(mp.proveedor),
            _fmt(mp.ubicacion),
            _fmt(mp.estado),
            mp.fecha_entrada.strftime("%d/%m/%Y") if mp.fecha_entrada else "",
            mp.dias_en_fabrica_num if mp.dias_en_fabrica_num is not None else "",
            ESTATUS_LABEL.get(mp.estatus_cobro, ""),
            mp.meses_a_cobrar if mp.meses_a_cobrar else "",
            _fmt(mp.observaciones),
        ])

    _apply_rows(ws)
    _autowidth(ws)

    fecha = timezone.localdate().strftime("%Y%m%d")
    return _make_response(wb, f"inventario_mp_{fecha}.xlsx")


# ── 2. Movimientos de Materia Prima ───────────────────────────────────────────

@login_required
def reporte_movimientos_mp(request):
    qs = MovimientoMP.objects.select_related("mp", "mp__cliente", "usuario").all()

    fecha_inicio    = request.GET.get("fecha_inicio")
    fecha_fin       = request.GET.get("fecha_fin")
    tipo_movimiento = request.GET.get("tipo_movimiento")
    cliente_id      = request.GET.get("cliente")

    if fecha_inicio:
        qs = qs.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__date__lte=fecha_fin)
    if tipo_movimiento:
        qs = qs.filter(tipo_movimiento=tipo_movimiento)
    if cliente_id:
        qs = qs.filter(mp__cliente_id=cliente_id)

    qs = qs.order_by("-fecha")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos MP"

    headers = [
        "Fecha", "N° MP", "Cliente", "Tipo Movimiento", "Peso (kg)",
        "Ubicación Origen", "Ubicación Destino", "Usuario", "Observaciones",
    ]
    _apply_header(ws, headers)

    for mov in qs:
        ws.append([
            timezone.localtime(mov.fecha).strftime("%d/%m/%Y %H:%M"),
            _fmt(mov.mp.numero_mp),
            _fmt(mov.mp.cliente.nombre if mov.mp.cliente else ""),
            _fmt(mov.get_tipo_movimiento_display()),
            float(mov.peso),
            _fmt(mov.ubicacion_origen),
            _fmt(mov.ubicacion_destino),
            _fmt(mov.usuario.username if mov.usuario else ""),
            _fmt(mov.observaciones),
        ])

    _apply_rows(ws)
    _autowidth(ws)

    fecha = timezone.localdate().strftime("%Y%m%d")
    return _make_response(wb, f"movimientos_mp_{fecha}.xlsx")


# ── 3. Órdenes de Producción ──────────────────────────────────────────────────

@login_required
def reporte_ordenes_produccion(request):
    qs = OrdenProduccion.objects.select_related(
        "cliente", "mp", "linea", "operador"
    ).all()

    fecha_inicio  = request.GET.get("fecha_inicio")
    fecha_fin     = request.GET.get("fecha_fin")
    estado        = request.GET.get("estado")
    tipo_proceso  = request.GET.get("tipo_proceso")
    cliente_id    = request.GET.get("cliente")

    if fecha_inicio:
        qs = qs.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__lte=fecha_fin)
    if estado:
        qs = qs.filter(estado=estado)
    if tipo_proceso:
        qs = qs.filter(tipo_proceso=tipo_proceso)
    if cliente_id:
        qs = qs.filter(cliente_id=cliente_id)

    qs = qs.order_by("-fecha", "folio_orden")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ordenes Produccion"

    headers = [
        "Folio", "Fecha", "Tipo Proceso", "Cliente", "N° MP",
        "Línea", "Operador", "Turno", "Prioridad", "Estado",
        "Hora Inicio", "Hora Fin",
        "T. Preparación (min)", "T. Proceso (min)", "T. Muerto (min)",
        "Peso Usado (kg)", "Peso Producido (kg)", "Scrap (kg)",
        "Merma (kg)", "Rendimiento (%)",
        "Paquetes", "Piezas", "Observaciones",
    ]
    _apply_header(ws, headers)

    for op in qs:
        ws.append([
            _fmt(op.folio_orden),
            op.fecha.strftime("%d/%m/%Y") if op.fecha else "",
            _fmt(op.get_tipo_proceso_display()),
            _fmt(op.cliente.nombre if op.cliente else ""),
            _fmt(op.mp.numero_mp if op.mp else ""),
            _fmt(op.linea.nombre if op.linea else ""),
            _fmt(op.operador.nombre if op.operador else ""),
            _fmt(op.get_turno_display() if op.turno else ""),
            _fmt(op.get_prioridad_display()),
            _fmt(op.get_estado_display()),
            op.hora_inicio.strftime("%H:%M") if op.hora_inicio else "",
            op.hora_fin.strftime("%H:%M") if op.hora_fin else "",
            op.tiempo_preparacion_min or "",
            op.tiempo_proceso_min or "",
            op.tiempo_muerto_min or "",
            float(op.peso_usado) if op.peso_usado else "",
            float(op.peso_producido) if op.peso_producido else "",
            float(op.scrap_total) if op.scrap_total else "",
            float(op.merma_kg) if op.merma_kg else "",
            float(op.rendimiento_porcentaje) if op.rendimiento_porcentaje else "",
            op.cantidad_paquetes or "",
            op.cantidad_piezas or "",
            _fmt(op.observaciones),
        ])

    _apply_rows(ws)
    _autowidth(ws)

    fecha = timezone.localdate().strftime("%Y%m%d")
    return _make_response(wb, f"ordenes_produccion_{fecha}.xlsx")


# ── 4. Detalles Slitter ───────────────────────────────────────────────────────

@login_required
def reporte_detalles_slitter(request):
    qs = DetalleSlitter.objects.select_related(
        "orden", "orden__cliente", "orden__mp"
    ).all()

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin    = request.GET.get("fecha_fin")
    cliente_id   = request.GET.get("cliente")

    if fecha_inicio:
        qs = qs.filter(orden__fecha__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(orden__fecha__lte=fecha_fin)
    if cliente_id:
        qs = qs.filter(orden__cliente_id=cliente_id)

    qs = qs.order_by("-orden__fecha", "orden__folio_orden", "no_corte")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalles Slitter"

    headers = [
        "Folio Orden", "Fecha Orden", "Cliente", "N° MP",
        "N° Corte", "Ancho (mm)", "Espesor (mm)", "Rebaba",
        "Peso (kg)", "Camber", "Material OK", "Observaciones",
    ]
    _apply_header(ws, headers)

    for d in qs:
        ws.append([
            _fmt(d.orden.folio_orden),
            d.orden.fecha.strftime("%d/%m/%Y") if d.orden.fecha else "",
            _fmt(d.orden.cliente.nombre if d.orden.cliente else ""),
            _fmt(d.orden.mp.numero_mp if d.orden.mp else ""),
            d.no_corte,
            float(d.ancho) if d.ancho else "",
            float(d.espesor) if d.espesor else "",
            _fmt(d.rebaba),
            float(d.peso) if d.peso else "",
            _fmt(d.camber),
            "Sí" if d.material_ok else "No",
            _fmt(d.observaciones),
        ])

    _apply_rows(ws)
    _autowidth(ws)

    fecha = timezone.localdate().strftime("%Y%m%d")
    return _make_response(wb, f"detalles_slitter_{fecha}.xlsx")


# ── 5. Producto Terminado ─────────────────────────────────────────────────────

@login_required
def reporte_producto_terminado(request):
    qs = ProductoTerminado.objects.select_related("cliente", "orden").all()

    estado       = request.GET.get("estado")
    cliente_id   = request.GET.get("cliente")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin    = request.GET.get("fecha_fin")

    if estado:
        qs = qs.filter(estado=estado)
    if cliente_id:
        qs = qs.filter(cliente_id=cliente_id)
    if fecha_inicio:
        qs = qs.filter(fecha_ingreso__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha_ingreso__date__lte=fecha_fin)

    qs = qs.order_by("-fecha_ingreso")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Producto Terminado"

    headers = [
        "N° PT", "Folio Orden", "Cliente", "Tipo Proceso",
        "Peso (kg)", "Paquetes", "Piezas",
        "Fecha Ingreso", "Estado", "Ubicación", "Observaciones",
    ]
    _apply_header(ws, headers)

    for pt in qs:
        ws.append([
            _fmt(pt.numero_pt),
            _fmt(pt.orden.folio_orden if pt.orden else ""),
            _fmt(pt.cliente.nombre if pt.cliente else ""),
            _fmt(pt.tipo_proceso),
            float(pt.peso_kg),
            pt.cantidad_paquetes or "",
            pt.cantidad_piezas or "",
            timezone.localtime(pt.fecha_ingreso).strftime("%d/%m/%Y %H:%M"),
            _fmt(pt.get_estado_display()),
            _fmt(pt.ubicacion),
            _fmt(pt.observaciones),
        ])

    _apply_rows(ws)
    _autowidth(ws)

    fecha = timezone.localdate().strftime("%Y%m%d")
    return _make_response(wb, f"producto_terminado_{fecha}.xlsx")


# ── 6. Clientes ───────────────────────────────────────────────────────────────

@login_required
def reporte_clientes(request):
    activo = request.GET.get("activo", "")
    qs = Cliente.objects.all().order_by("nombre")
    if activo == "1":
        qs = qs.filter(activo=True)
    elif activo == "0":
        qs = qs.filter(activo=False)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = ["Código Cliente", "Nombre", "Activo"]
    _apply_header(ws, headers)

    for c in qs:
        ws.append([
            _fmt(c.codigo_cliente),
            _fmt(c.nombre),
            "Sí" if c.activo else "No",
        ])

    _apply_rows(ws)
    _autowidth(ws)

    fecha = timezone.localdate().strftime("%Y%m%d")
    return _make_response(wb, f"clientes_{fecha}.xlsx")
