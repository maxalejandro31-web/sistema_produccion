import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db.models import Sum, Max, ProtectedError
from django.utils import timezone

from .forms import MateriaPrimaForm, ClienteForm, RegistrarMovimientoForm
from .models import MateriaPrima, Cliente, MovimientoMP
from produccion.models import OrdenProduccion, DetalleSlitter
from dashboard.models import registrar_historial
from core.decorators import solo_dueno_puede_eliminar


@login_required
def captura_mp(request):
    if request.method == 'POST':
        form = MateriaPrimaForm(request.POST, request.FILES)
        if form.is_valid():
            mp_nueva = form.save()
            registrar_historial(request, 'MateriaPrima', mp_nueva.id, str(mp_nueva), 'CREAR', f'MP {mp_nueva.numero_mp} registrada.')
            messages.success(request, 'Materia prima registrada correctamente.')
            form = MateriaPrimaForm()
    else:
        form = MateriaPrimaForm()

    return render(request, 'inventario/captura_mp.html', {'form': form})


@login_required
def lista_mp(request):
    busqueda      = request.GET.get('q', '')
    tipo          = request.GET.get('tipo', '')
    estado        = request.GET.get('estado', '')
    fecha_inicio  = request.GET.get('fecha_inicio', '')
    fecha_fin     = request.GET.get('fecha_fin', '')
    cliente_id    = request.GET.get('cliente', '')
    cobro         = request.GET.get('cobro', '')

    hoy = timezone.localdate()

    qs = MateriaPrima.objects.select_related('cliente').order_by('-id')

    if busqueda:
        qs = qs.filter(numero_mp__icontains=busqueda)
    if tipo:
        qs = qs.filter(tipo_mp=tipo)
    if estado:
        qs = qs.filter(estado=estado)
    if fecha_inicio:
        qs = qs.filter(fecha_entrada__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha_entrada__lte=fecha_fin)
    if cliente_id:
        qs = qs.filter(cliente_id=cliente_id)
    if cobro == 'vencido':
        qs = qs.filter(fecha_entrada__lt=hoy - datetime.timedelta(days=30))
    elif cobro == 'por_vencer':
        qs = qs.filter(
            fecha_entrada__range=(hoy - datetime.timedelta(days=30), hoy - datetime.timedelta(days=23))
        )
    elif cobro == 'libre':
        qs = qs.filter(fecha_entrada__gte=hoy - datetime.timedelta(days=22))

    mp_vencidas_count   = MateriaPrima.objects.filter(fecha_entrada__lt=hoy - datetime.timedelta(days=30)).exclude(cliente__nombre='MAQUILAS Y SERVICIOS JC').count()
    mp_por_vencer_count = MateriaPrima.objects.filter(
        fecha_entrada__range=(hoy - datetime.timedelta(days=30), hoy - datetime.timedelta(days=23))
    ).exclude(cliente__nombre='MAQUILAS Y SERVICIOS JC').count()

    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'inventario/lista_mp.html', {
        'materias_primas': page_obj,
        'page_obj': page_obj,
        'busqueda': busqueda,
        'tipo': tipo,
        'estado': estado,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'cliente_id': cliente_id,
        'cobro': cobro,
        'mp_vencidas_count': mp_vencidas_count,
        'mp_por_vencer_count': mp_por_vencer_count,
        'clientes': Cliente.objects.filter(activo=True).order_by('nombre'),
    })


@solo_dueno_puede_eliminar
@require_POST
def eliminar_mp(request, mp_id):
    mp = get_object_or_404(MateriaPrima, id=mp_id)
    numero_mp = mp.numero_mp
    mp_id_original = mp.id
    mp_str = str(mp)

    try:
        mp.delete()
    except ProtectedError:
        messages.error(
            request,
            f'No se puede eliminar la MP {numero_mp} porque ya tiene órdenes de '
            'producción asociadas. Primero hay que eliminar o reasignar esas '
            'órdenes antes de borrar la materia prima (esto también borraría '
            'su historial de movimientos).'
        )
        return redirect('detalle_mp', mp_id=mp_id)

    registrar_historial(request, 'MateriaPrima', mp_id_original, mp_str, 'ELIMINAR', f'MP {numero_mp} eliminada.')
    messages.success(request, f'Materia prima {numero_mp} eliminada correctamente.')
    return redirect('lista_mp')


@login_required
def editar_mp(request, mp_id):
    if not (
        request.user.is_superuser or
        request.user.groups.filter(name__in=['Administrador', 'Supervisor', 'Coordinador']).exists()
    ):
        return HttpResponse("No tienes permiso para editar materia prima.")

    mp = get_object_or_404(MateriaPrima, id=mp_id)

    if request.method == 'POST':
        form = MateriaPrimaForm(request.POST, request.FILES, instance=mp)
        if form.is_valid():
            form.save()
            registrar_historial(request, 'MateriaPrima', mp.id, str(mp), 'EDITAR', f'MP {mp.numero_mp} actualizada.')
            messages.success(request, f'Materia prima {mp.numero_mp} actualizada correctamente.')
            return redirect('lista_mp')
    else:
        form = MateriaPrimaForm(instance=mp)

    pdf_url = None
    if mp.archivo_pdf:
        try:
            pdf_url = mp.archivo_pdf.url
        except Exception:
            pass

    return render(request, 'inventario/editar_mp.html', {
        'form': form,
        'mp': mp,
        'pdf_url': pdf_url,
    })


@login_required
def detalle_mp(request, mp_id):
    if not (
        request.user.is_superuser or
        request.user.groups.filter(name__in=['Administrador', 'Supervisor', 'Operador', 'Almacen', 'Coordinador']).exists()
    ):
        return HttpResponse("No tienes permiso para ver la materia prima.")

    mp = get_object_or_404(MateriaPrima, id=mp_id)

    ordenes_relacionadas = OrdenProduccion.objects.select_related(
        'cliente', 'linea'
    ).filter(mp_id=mp.id).order_by('-id')

    resumen = ordenes_relacionadas.aggregate(
        total_consumido=Sum('peso_usado'),
        total_producido=Sum('peso_producido'),
        total_scrap=Sum('scrap_total'),
    )

    total_consumido = resumen['total_consumido'] or 0
    total_producido = resumen['total_producido'] or 0
    total_scrap = resumen['total_scrap'] or 0
    cantidad_ordenes = ordenes_relacionadas.count()

    movimientos = MovimientoMP.objects.filter(mp=mp).select_related('usuario').order_by('-fecha')

    from dashboard.models import HistorialCambio
    historial = HistorialCambio.objects.filter(
        tipo_objeto='MateriaPrima', objeto_id=mp_id
    ).select_related('usuario')

    pdf_url = None
    if mp.archivo_pdf:
        try:
            pdf_url = mp.archivo_pdf.url
        except Exception:
            pass

    return render(request, 'inventario/detalle_mp.html', {
        'mp': mp,
        'ordenes_relacionadas': ordenes_relacionadas,
        'total_consumido': total_consumido,
        'total_producido': total_producido,
        'total_scrap': total_scrap,
        'cantidad_ordenes': cantidad_ordenes,
        'movimientos': movimientos,
        'historial': historial,
        'pdf_url': pdf_url,
    })


def _datos_reporte_rollo(mp):
    """Reconstruye todo el árbol de un rollo: cada orden de slitter que lo
    cortó -> sus cortes -> la orden de fleje (si ya existe) que consumió
    cada corte -> sus descargas -- junto con los totales generales del
    rollo. Es el mismo dato para la vista en pantalla y para el Excel, para
    que ambos digan siempre lo mismo."""
    from produccion.models import OrdenProduccion
    from produccion.views import _resumen_aprovechamiento_mp

    ordenes_slitter = OrdenProduccion.objects.filter(
        mp=mp, tipo_proceso='slitter'
    ).select_related('cliente', 'linea').prefetch_related('detalles_slitter').order_by('fecha', 'id')

    bloques = []
    total_scrap_slitter = 0.0
    total_descarte_slitter = 0.0
    total_scrap_fleje = 0.0
    total_fleje_producido = 0.0

    for orden in ordenes_slitter:
        detalles = list(orden.detalles_slitter.all())
        ordenes_fleje_hijas = OrdenProduccion.objects.filter(
            tipo_proceso='fleje',
            pt_origen__detalle_slitter__orden=orden,
        ).select_related(
            'pt_origen', 'pt_origen__detalle_slitter'
        ).prefetch_related('detalles_fleje').order_by('fecha', 'id')

        resumen = _resumen_aprovechamiento_mp(orden, detalles, ordenes_fleje_hijas)

        bloques.append({
            'orden': orden,
            'detalles': detalles,
            'ordenes_fleje_hijas': ordenes_fleje_hijas,
            'resumen': resumen,
        })

        total_scrap_slitter += resumen['peso_scrap_slitter']
        total_descarte_slitter += resumen['peso_descarte_slitter']
        total_scrap_fleje += resumen['scrap_fleje']
        total_fleje_producido += resumen['peso_fleje_producido']

    peso_rollo = float(mp.peso) if mp.peso else 0.0
    total_scrap_general = total_scrap_slitter + total_descarte_slitter + total_scrap_fleje
    aprovechamiento_pct = round((total_fleje_producido / peso_rollo) * 100, 2) if peso_rollo else None

    return {
        'bloques': bloques,
        'peso_rollo': peso_rollo,
        'total_scrap_slitter': total_scrap_slitter,
        'total_descarte_slitter': total_descarte_slitter,
        'total_scrap_fleje': total_scrap_fleje,
        'total_fleje_producido': total_fleje_producido,
        'total_scrap_general': total_scrap_general,
        'aprovechamiento_pct': aprovechamiento_pct,
    }


@login_required
def reporte_rollo(request, mp_id):
    """Reporte completo de trazabilidad de un rollo: MP -> cortes de
    slitter -> descargas de fleje de cada corte -> totales. Junta en una
    sola vista todas las órdenes de slitter que se hayan hecho sobre este
    rollo (un rollo se puede cortar en más de una orden, p. ej. un
    remanente que se vuelve a pasar después)."""
    mp = get_object_or_404(MateriaPrima, id=mp_id)
    datos = _datos_reporte_rollo(mp)
    return render(request, 'inventario/reporte_rollo.html', {'mp': mp, **datos})


@login_required
def reporte_rollo_excel(request, mp_id):
    """Mismo reporte que reporte_rollo, pero exportado a un .xlsx con el
    mismo formato que se llevaba a mano en planta (No. de rollo / peso /
    calibre / ancho, tabla de SLITTER, una tabla de FLEJES por cada corte
    ya procesado, y los totales al final)."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse as _HttpResponse

    mp = get_object_or_404(MateriaPrima, id=mp_id)
    datos = _datos_reporte_rollo(mp)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (mp.numero_mp or f'MP-{mp.id}')[:31]

    negrita = Font(bold=True)
    titulo_fill = PatternFill('solid', fgColor='1F2D3D')
    titulo_font = Font(bold=True, color='FFFFFF')
    centrado = Alignment(horizontal='center')

    fila = 1

    def escribir_titulo(texto, ncols=7):
        nonlocal fila
        ws.cell(row=fila, column=1, value=texto)
        ws.cell(row=fila, column=1).font = titulo_font
        ws.cell(row=fila, column=1).fill = titulo_fill
        for c in range(2, ncols + 1):
            ws.cell(row=fila, column=c).fill = titulo_fill
        fila += 1

    escribir_titulo('No. DE ROLLO')
    ws.cell(row=fila, column=1, value=mp.numero_mp).font = negrita
    fila += 1

    ws.cell(row=fila, column=1, value='PESO DE ROLLO').font = negrita
    ws.cell(row=fila, column=2, value='ESPESOR (MILS)').font = negrita
    ws.cell(row=fila, column=3, value='ANCHO').font = negrita
    fila += 1
    ws.cell(row=fila, column=1, value=float(mp.peso) if mp.peso else None)
    ws.cell(row=fila, column=2, value=float(mp.espesor_mils) if mp.espesor_mils else None)
    ws.cell(row=fila, column=3, value=float(mp.ancho) if mp.ancho else None)
    fila += 2

    for bloque in datos['bloques']:
        orden = bloque['orden']
        escribir_titulo(f"SLITTER — Orden {orden.folio_orden or orden.id} · {orden.fecha or ''}")
        headers = ['No. de corte', 'Ancho', 'Espesor', 'Rebaba', 'Peso', 'Clasificación', 'Peso scrap/descarte']
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=fila, column=i, value=h)
            c.font = negrita
        fila += 1
        for d in bloque['detalles']:
            ws.cell(row=fila, column=1, value=d.folio_corte or d.no_corte)
            ws.cell(row=fila, column=2, value=float(d.ancho) if d.ancho else None)
            ws.cell(row=fila, column=3, value=float(d.espesor) if d.espesor else None)
            ws.cell(row=fila, column=4, value=d.rebaba or '')
            ws.cell(row=fila, column=5, value=float(d.peso) if d.peso else None)
            ws.cell(row=fila, column=6, value=d.get_clasificacion_display())
            ws.cell(row=fila, column=7, value=float(d.peso_merma) if d.peso_merma else None)
            fila += 1
        fila += 1

        for oh in bloque['ordenes_fleje_hijas']:
            origen = oh.pt_origen.detalle_slitter.folio_corte if (oh.pt_origen and oh.pt_origen.detalle_slitter) else str(oh.pt_origen)
            escribir_titulo(f"FLEJES — Orden {oh.folio_orden or oh.id} · origen {origen} · {oh.fecha or ''} · {oh.tipo_fleje or ''}")
            headers = ['No. descarga', 'Tipo de fleje', 'Peso descarga', '# Flejes', 'Peso x fleje', 'Ancho', 'Folio descarga']
            for i, h in enumerate(headers, start=1):
                c = ws.cell(row=fila, column=i, value=h)
                c.font = negrita
            fila += 1
            for d in oh.detalles_fleje.all():
                ws.cell(row=fila, column=1, value=d.numero_descarga or d.no_fleje)
                ws.cell(row=fila, column=2, value=oh.tipo_fleje or '')
                ws.cell(row=fila, column=3, value=float(d.peso_descarga) if d.peso_descarga else None)
                ws.cell(row=fila, column=4, value=d.numero_flejes)
                ws.cell(row=fila, column=5, value=d.peso_por_fleje)
                ws.cell(row=fila, column=6, value=float(d.ancho) if d.ancho else None)
                ws.cell(row=fila, column=7, value=d.folio_descarga or '')
                fila += 1
            ws.cell(row=fila, column=1, value='Peso total').font = negrita
            ws.cell(row=fila, column=3, value=float(oh.peso_producido) if oh.peso_producido else None).font = negrita
            ws.cell(row=fila, column=4, value=f"Scrap: {oh.scrap_total or 0} kg").font = negrita
            fila += 2

    fila += 1
    ws.cell(row=fila, column=1, value='TOTAL DESCARGAS (fleje)').font = negrita
    ws.cell(row=fila, column=2, value='TOTAL SCRAP SLITTER').font = negrita
    ws.cell(row=fila, column=3, value='TOTAL DESCARTE SLITTER').font = negrita
    ws.cell(row=fila, column=4, value='TOTAL SCRAP FLEJE').font = negrita
    ws.cell(row=fila, column=5, value='TOTAL SCRAP (todo)').font = negrita
    ws.cell(row=fila, column=6, value='TOTAL ROLLO').font = negrita
    ws.cell(row=fila, column=7, value='% APROVECHAMIENTO').font = negrita
    fila += 1
    ws.cell(row=fila, column=1, value=round(datos['total_fleje_producido'], 2))
    ws.cell(row=fila, column=2, value=round(datos['total_scrap_slitter'], 2))
    ws.cell(row=fila, column=3, value=round(datos['total_descarte_slitter'], 2))
    ws.cell(row=fila, column=4, value=round(datos['total_scrap_fleje'], 2))
    ws.cell(row=fila, column=5, value=round(datos['total_scrap_general'], 2))
    ws.cell(row=fila, column=6, value=round(datos['peso_rollo'], 2))
    ws.cell(row=fila, column=7, value=datos['aprovechamiento_pct'])

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 18

    response = _HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = f"Reporte_{mp.numero_mp or mp.id}.xlsx".replace('/', '-')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


@login_required
def lista_clientes(request):
    if not (
        request.user.is_superuser or
        request.user.groups.filter(name__in=['Administrador', 'Supervisor', 'Coordinador']).exists()
    ):
        return HttpResponse("No tienes permiso para ver clientes.")

    busqueda = request.GET.get('q', '')
    qs = Cliente.objects.all().order_by('nombre')
    if busqueda:
        qs = qs.filter(nombre__icontains=busqueda)

    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'inventario/lista_clientes.html', {
        'clientes': page_obj,
        'page_obj': page_obj,
        'busqueda': busqueda,
    })


@login_required
def captura_cliente(request):
    if not (
        request.user.is_superuser or
        request.user.groups.filter(name__in=['Administrador', 'Supervisor', 'Coordinador']).exists()
    ):
        return HttpResponse("No tienes permiso para capturar clientes.")

    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente registrado correctamente.')
            form = ClienteForm()
    else:
        form = ClienteForm()

    return render(request, 'inventario/captura_cliente.html', {'form': form})


@login_required
def editar_cliente(request, cliente_id):
    if not (
        request.user.is_superuser or
        request.user.groups.filter(name__in=['Administrador', 'Supervisor', 'Coordinador']).exists()
    ):
        return HttpResponse("No tienes permiso para editar clientes.")

    cliente = get_object_or_404(Cliente, id=cliente_id)

    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, f'Cliente {cliente.nombre} actualizado correctamente.')
            return redirect('lista_clientes')
    else:
        form = ClienteForm(instance=cliente)

    return render(request, 'inventario/editar_cliente.html', {
        'form': form,
        'cliente': cliente,
    })


@login_required
def registrar_movimiento(request, mp_id):
    if not (
        request.user.is_superuser or
        request.user.groups.filter(name__in=['Administrador', 'Supervisor', 'Almacen', 'Coordinador']).exists()
    ):
        messages.error(request, 'No tienes permiso para registrar movimientos.')
        return redirect('detalle_mp', mp_id=mp_id)

    mp = get_object_or_404(MateriaPrima, id=mp_id)

    if request.method == 'POST':
        form = RegistrarMovimientoForm(request.POST)
        if form.is_valid():
            mov = form.save(commit=False)
            mov.mp = mp
            mov.usuario = request.user
            mov.save()
            etiquetas = {
                'ENTRADA': 'Entrada', 'CONSUMO': 'Consumo',
                'AJUSTE_POSITIVO': 'Ajuste positivo', 'AJUSTE_NEGATIVO': 'Ajuste negativo',
                'MERMA': 'Merma', 'TRASPASO': 'Traspaso', 'SALIDA': 'Salida',
            }
            tipo_label = etiquetas.get(mov.tipo_movimiento, mov.tipo_movimiento)
            registrar_historial(request, 'MateriaPrima', mp.id, str(mp), 'MOVIMIENTO',
                f'{tipo_label} de {mov.peso} kg en MP {mp.numero_mp}.')
            messages.success(request, f'Movimiento "{tipo_label}" de {mov.peso} kg registrado correctamente.')
            return redirect('detalle_mp', mp_id=mp.id)
    else:
        form = RegistrarMovimientoForm()

    return render(request, 'inventario/registrar_movimiento.html', {
        'mp': mp,
        'form': form,
    })


@login_required
def dar_salida_mp(request, mp_id):
    if not (
        request.user.is_superuser or
        request.user.groups.filter(name__in=['Administrador', 'Supervisor', 'Almacen', 'Coordinador']).exists()
    ):
        messages.error(request, 'No tienes permiso para registrar salidas.')
        return redirect('detalle_mp', mp_id=mp_id)

    mp = get_object_or_404(MateriaPrima, id=mp_id)
    clientes = Cliente.objects.filter(activo=True).order_by('nombre')

    if request.method == 'POST':
        peso_str      = request.POST.get('peso', '').replace(',', '.')
        cliente_id    = request.POST.get('cliente_id') or None
        fecha_salida  = request.POST.get('fecha_salida')
        observaciones = request.POST.get('observaciones', '')

        try:
            peso = float(peso_str)
            if peso <= 0:
                raise ValueError
        except (ValueError, TypeError):
            messages.error(request, 'El peso debe ser un número positivo.')
            return render(request, 'inventario/dar_salida_mp.html', {'mp': mp, 'clientes': clientes})

        if mp.peso_restante is not None and peso > float(mp.peso_restante):
            messages.error(request, f'El peso ingresado ({peso} kg) supera el peso restante ({mp.peso_restante} kg).')
            return render(request, 'inventario/dar_salida_mp.html', {'mp': mp, 'clientes': clientes})

        cliente_nombre = ''
        if cliente_id:
            try:
                cliente_nombre = Cliente.objects.get(pk=cliente_id).nombre
            except Cliente.DoesNotExist:
                pass

        from django.utils import timezone as tz
        fecha_dt = tz.now()
        if fecha_salida:
            try:
                import datetime
                fecha_dt = datetime.datetime.fromisoformat(fecha_salida)
                if timezone.is_naive(fecha_dt):
                    fecha_dt = timezone.make_aware(fecha_dt)
            except Exception:
                pass

        MovimientoMP.objects.create(
            mp=mp,
            tipo_movimiento='SALIDA',
            peso=peso,
            fecha=fecha_dt,
            ubicacion_origen=mp.ubicacion or '',
            ubicacion_destino=cliente_nombre,
            observaciones=observaciones,
            usuario=request.user,
        )

        registrar_historial(request, 'MateriaPrima', mp.id, str(mp), 'MOVIMIENTO',
            f'Salida de {peso} kg de MP {mp.numero_mp} hacia {cliente_nombre or "destino no especificado"}.')
        messages.success(request, f'Salida de {peso} kg registrada. Peso restante: {mp.peso_restante} kg.')
        return redirect('detalle_mp', mp_id=mp.id)

    return render(request, 'inventario/dar_salida_mp.html', {'mp': mp, 'clientes': clientes})


@login_required
def lista_salidas_mp(request):
    busqueda     = request.GET.get('q', '')
    cliente_id   = request.GET.get('cliente', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin    = request.GET.get('fecha_fin', '')

    qs = MovimientoMP.objects.filter(tipo_movimiento='SALIDA').select_related(
        'mp', 'mp__cliente', 'usuario'
    ).order_by('-fecha')

    if busqueda:
        qs = qs.filter(mp__numero_mp__icontains=busqueda)
    if cliente_id:
        qs = qs.filter(mp__cliente_id=cliente_id)
    if fecha_inicio:
        qs = qs.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__date__lte=fecha_fin)

    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'inventario/lista_salidas_mp.html', {
        'salidas': page_obj,
        'page_obj': page_obj,
        'busqueda': busqueda,
        'cliente_id': cliente_id,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'clientes': Cliente.objects.filter(activo=True).order_by('nombre'),
    })


@login_required
def api_datos_mp(request, mp_id):
    mp = get_object_or_404(MateriaPrima, id=mp_id)
    # Un mismo rollo puede pasar por la slitter varias veces en órdenes
    # distintas; el número de corte debe seguir la secuencia del rollo
    # completo, no reiniciar en cada orden nueva.
    ultimo_corte = DetalleSlitter.objects.filter(orden__mp=mp).aggregate(Max('no_corte'))['no_corte__max']
    return JsonResponse({
        'cliente_id': mp.cliente_id,
        'cliente_nombre': str(mp.cliente) if mp.cliente else '',
        'material': mp.material or '',
        'espesor_valor': str(mp.espesor_valor) if mp.espesor_valor else '',
        'unidad_espesor': mp.unidad_espesor or '',
        'espesor_mils': str(mp.espesor_mils) if mp.espesor_mils else '',
        'ancho': str(mp.ancho) if mp.ancho else '',
        'peso_restante': str(mp.peso_restante) if mp.peso_restante else '',
        'ubicacion': mp.ubicacion or '',
        'proximo_no_corte': (ultimo_corte or 0) + 1,
    })
